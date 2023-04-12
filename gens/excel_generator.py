import os
import re
import requests
from datetime import date
from PIL import Image as Img
from openpyxl import Workbook, load_workbook
from gens.excel_styles import fill_black, ft_arial_black, ft_arial_white, center_align, cats_align, ft_arial_blue, ft_calibri_black, urls_align
from openpyxl.drawing.image import Image


def clear(a: str):
    return re.sub(r'[\?\\\/\*\|\"<>]', '', a)


def download_img(url, art):
    form = url.split('.')[-1]

    resp = requests.get(url)
    with open(f'excel_images\\source\\pic_{clear(art)}.{form}', 'wb') as file:
        file.write(resp.content)


def resize_image(sizes, art, form):
    img = Img.open(f'excel_images\\source\\pic_{clear(art)}.{form}')
    new_img = img.resize(sizes)
    new_img.save(f'excel_images\\small\\pic_{clear(art)}_small.{form}')
    os.remove(f'excel_images\\source\\pic_{clear(art)}.{form}')


def up_round(i, ri):
    if i % ri > 0:
        return int((i // ri)*ri+ri)
    else:
        return int((i // ri)*ri)


def create_price(price_type, cyn, margin, src_price, ri):
    if price_type == 1: # проценты
        price = up_round(float(src_price)*cyn+float(src_price)*cyn*margin/100, ri)
    elif price_type == 2: # маржа
        price = up_round(float(src_price)*cyn + margin, ri)
    else:
        print('Error: Incorrect price type.')
        price.value = 0
    return price


def create_rozn_table(r_text):
    wb = Workbook()
    ws = wb.active

    ws.merge_cells('A1:D1')
    ws.merge_cells('E1:F1')

    file_name = f'output\\files\\rozn-{date.today().strftime("%d%m%Y")}.xlsx'

    # wb = load_workbook(file_name)
    # ws = wb['test']

    ws.row_dimensions[1].height = 86.25

    ws.column_dimensions['A'].width = 11.86+0.72
    ws.column_dimensions['B'].width = 12.43+0.72
    ws.column_dimensions['C'].width = 11.29+0.72
    ws.column_dimensions['D'].width = 18.71+0.72
    ws.column_dimensions['E'].width = 10.43+0.72
    ws.column_dimensions['F'].width = 69.43+0.72
    ws.column_dimensions['G'].width = 20+0.72
    ws.column_dimensions['H'].width = 11.86+0.72

    a1 = ws['A1']

    a1.font = ft_arial_blue
    a1.alignment = center_align
    a1.value = r_text
    a1.hyperlink = 'https://exclusive-only.ru/'

    im_file = os.listdir("input\\picrozn")[0]
    image = Img.open(f'input\\picrozn\\{im_file}')
    new_img = image.resize((115, 115,))
    new_img.save(r'excel_images\exonly.png')

    img = Image(r'excel_images\exonly.png')
    ws.add_image(img, 'G1')

    a2 = ws['A2']

    a2.font = ft_arial_white
    a2.fill = fill_black
    a2.alignment = cats_align
    a2.value = 'Размер EU'

    b2 = ws['B2']

    b2.font = ft_arial_white
    b2.fill = fill_black
    b2.alignment = cats_align
    b2.value = 'Изображение'

    c2 = ws['C2']

    c2.font = ft_arial_white
    c2.fill = fill_black
    c2.alignment = cats_align
    c2.value = 'Цена'

    d2 = ws['D2']

    d2.font = ft_arial_white
    d2.fill = fill_black
    d2.alignment = cats_align
    d2.value = 'Ссылка'

    e2 = ws['E2']

    e2.font = ft_arial_white
    e2.fill = fill_black
    e2.alignment = cats_align
    e2.value = 'Бренд'

    f2 = ws['F2']

    f2.font = ft_arial_white
    f2.fill = fill_black
    f2.alignment = cats_align
    f2.value = 'Наименование'

    g2 = ws['G2']

    g2.font = ft_arial_white
    g2.fill = fill_black
    g2.alignment = cats_align
    g2.value = 'Код артикула'

    ws.freeze_panes = 'A3'

    wb.save(file_name)

    return file_name


def create_opt_table(o_text, flag):
    abc = 'ABCDEFGHI'
    wb = Workbook()
    ws = wb.active

    file_name = f'output\\files\\opt-{date.today().strftime("%d%m%Y")}.xlsx'

    ws.merge_cells('A1:E1')
    ws.merge_cells('F1:H1')

    ws.row_dimensions[1].height = 82.5

    ws.column_dimensions[f'{abc[0]}'].width = 11.86+0.72
    ws.column_dimensions[f'{abc[1]}'].width = 12.43+0.72
    ws.column_dimensions[f'{abc[2]}'].width = 11.29+0.72
    i = 3
    if flag == True:
        ws.column_dimensions[f'{abc[i]}'].width = 16.14+0.72
        i += 1
    ws.column_dimensions[f'{abc[i]}'].width = 18.71+0.72
    i += 1
    ws.column_dimensions[f'{abc[i]}'].width = 78.86+0.72
    i += 1
    ws.column_dimensions[f'{abc[i]}'].width = 20+0.72
    i += 1
    ws.column_dimensions[f'{abc[i]}'].width = 12.57+0.72
    i += 1
    ws.column_dimensions[f'{abc[i]}'].width = 11.86+0.72
    i += 1

    im_file = os.listdir("input\\picopt")[0]
    image = Img.open(f'input\\picopt\\{im_file}')
    new_img = image.resize((115, 115,))
    new_img.save(r'excel_images\opt.png')

    img = Image(r'excel_images\opt.png')
    ws.add_image(img, 'A1')

    a1 = ws['A1']

    a1.font = ft_arial_blue
    a1.alignment = center_align
    a1.value = o_text
    a1.hyperlink = 'https://t.me/+ZGTZ-NvQUwllMmI1'

    a2 = ws['A2']

    a2.font = ft_arial_white
    a2.fill = fill_black
    a2.alignment = cats_align
    a2.value = 'Размер EU'

    b2 = ws['B2']

    b2.font = ft_arial_white
    b2.fill = fill_black
    b2.alignment = cats_align
    b2.value = 'Изображение'

    c2 = ws['C2']

    c2.font = ft_arial_white
    c2.fill = fill_black
    c2.alignment = cats_align
    c2.value = 'Опт цена'
    i = 3
    if flag == True:
        d2 = ws[f'{abc[i]}2']

        d2.font = ft_arial_white
        d2.fill = fill_black
        d2.alignment = cats_align
        d2.value = 'Цена'
        i += 1
    
    e2 = ws[f'{abc[i]}2']

    e2.font = ft_arial_white
    e2.fill = fill_black
    e2.alignment = cats_align
    e2.value = 'Бренд'
    i += 1

    f2 = ws[f'{abc[i]}2']

    f2.font = ft_arial_white
    f2.fill = fill_black
    f2.alignment = cats_align
    f2.value = 'Наименование'
    i += 1

    g2 = ws[f'{abc[i]}2']

    g2.font = ft_arial_white
    g2.fill = fill_black
    g2.alignment = cats_align
    g2.value = 'Код артикула'
    i += 1

    h2 = ws[f'{abc[i]}2']

    h2.font = ft_arial_white
    h2.fill = fill_black
    h2.alignment = cats_align
    h2.value = 'Наличие'
    
    ws.freeze_panes = 'A3'

    wb.save(file_name)

    return file_name


def add_opt_row(i, file_name, row, opt_price_type, cny, opt_margin, opt_round_size, rozn_price_type, rozn_margin, rozn_round_size, nv, p, flag):
    abc = 'ABCDEFGHI'
    # Открыли табличку
    wb = load_workbook(file_name)
    # Выбрали лист
    ws = wb.active
    # Указали высоту строки, в которой мы будем работать
    ws.row_dimensions[i+3].height = 79.5
    j = 0
    # Присваеваем переменным объект ячейки для удобства
    sizeEU = ws[f'{abc[j]}{i+3}']
    j += 1
    optPrice = ws[f'{abc[j]}{i+3}']
    j += 1
    if flag == True:
        price = ws[f'{abc[j]}{i+3}']
        j += 1
    brand = ws[f'{abc[j]}{i+3}']
    j += 1
    name = ws[f'{abc[j]}{i+3}']
    j += 1
    article = ws[f'{abc[j]}{i+3}']
    j += 1
    availability = ws[f'{abc[j]}{i+3}']
    j += 1

    # Указываем стили для ячеек
    sizeEU.font = ft_arial_black
    sizeEU.alignment = center_align
    optPrice.font = ft_calibri_black
    optPrice.alignment = center_align
    if flag == True:
        price.font = ft_calibri_black
        price.alignment = center_align
    brand.font = ft_calibri_black
    brand.alignment = center_align
    name.font = ft_calibri_black
    name.alignment = center_align
    article.font = ft_calibri_black
    article.alignment = center_align
    availability.font = ft_arial_black
    availability.alignment = center_align

    # Заполняем ячейки
    sizeEU.value = row['Наименование артикула']
    brand.value = p.loc[p['ID товара'] == row['ID товара'], ['Бренд']].values.item(0)
    name.value = row['Наименование']
    article.value = row['Код артикула']
    availability.value = row['В наличии @Наличие в Москве']

    src_price = float(row['Цена'])
    article_code = row['Код артикула']
    optPrice.value = create_price(opt_price_type, cny, opt_margin, src_price, opt_round_size)
    if flag == True:
        price.value = create_price(rozn_price_type, cny, rozn_margin, src_price, rozn_round_size)
    img_url = p.loc[p['ID товара'] == row['ID товара'], ['Изображения товаров']].values.item(0)
    form = img_url.split('.')[-1]

    if os.path.exists(f'excel_images\\small\\pic_{clear(article_code)}_small.{form}') == False:
        download_img(img_url, article_code)
        resize_image((111, 84,), article_code, form)

    img = Image(f'excel_images\\small\\pic_{clear(article_code)}_small.{form}')

    ws.add_image(img, f'B{i+3}')
    if article_code in nv['Код артикула'].to_numpy():
        new_img = Image(r'excel_images\new.jpg')
        ws.add_image(new_img, f'{abc[j]}{i+3}')
    wb.save(file_name)


def add_rozn_row(i, file_name, row, price_type, cny, margin, round_size, nv, p):

    wb = load_workbook(file_name)
    ws = wb.active
    ws.row_dimensions[i+3].height = 79.5

    sizeEU = ws[f'A{i+3}']
    price = ws[f'C{i+3}']
    link = ws[f'D{i+3}']
    brand = ws[f'E{i+3}']
    name = ws[f'F{i+3}']
    article = ws[f'G{i+3}']

    sizeEU.font = ft_calibri_black
    sizeEU.alignment = center_align

    ws[f'B{i+3}'].alignment = center_align

    price.font = ft_calibri_black
    price.alignment = center_align

    link.font = ft_calibri_black
    link.alignment = urls_align

    brand.font = ft_calibri_black
    brand.alignment = center_align

    name.font = ft_calibri_black
    name.alignment = center_align

    article.font = ft_calibri_black
    article.alignment = center_align

    sizeEU.value = row['Наименование артикула']
    brand.value = p.loc[p['ID товара'] == row['ID товара'], ['Бренд']].values.item(0)
    name.value = row['Наименование']
    article.value = row['Код артикула']
    link.value = f'https://exclusive-only.ru/{row["Ссылка на витрину"]}'

    src_price = float(row['Цена'])
    art_code = row['Код артикула']
    price.value = create_price(price_type, cny, margin, src_price, round_size)
    img_url = p.loc[p['ID товара'] == row['ID товара'], ['Изображения товаров']].values.item(0)
    form = img_url.split('.')[-1]

    if os.path.exists(f'excel_images\\small\\pic_{clear(art_code)}_small.{form}') == False:
        download_img(img_url, art_code)
        resize_image((111, 84,), art_code, form)

    img = Image(f'excel_images\\small\\pic_{clear(art_code)}_small.{form}')

    ws.add_image(img, f'B{i+3}')
    if art_code in nv['Код артикула'].to_numpy():
        new_img = Image(r'excel_images\new.jpg')
        ws.add_image(new_img, f'H{i+3}')
    
    print(f'add excel row {i}')

    wb.save(file_name)


def main():
    # file_name = 'output\\files\\opt.xlsx'
    # create_opt_table(file_name)
    file_name = create_rozn_table()
    

if __name__ == '__main__':
    main()